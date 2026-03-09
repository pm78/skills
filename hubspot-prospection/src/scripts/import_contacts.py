#!/usr/bin/env python3
"""
HubSpot Contact Batch Import Script

Imports contacts from CSV/Excel files into HubSpot CRM via the Contacts API.
Handles column mapping, validation, deduplication, and batch creation.

Usage:
    # Import from CSV with auto-mapping
    python import_contacts.py --token TOKEN --file prospects.csv

    # Import with explicit column mapping
    python import_contacts.py --token TOKEN --file prospects.xlsx \
        --mapping "Prénom=firstname,Nom=lastname,Email=email,Société=company"

    # Import with lifecycle stage and dry run
    python import_contacts.py --token TOKEN --file prospects.csv \
        --lifecycle-stage marketingqualifiedlead --dry-run

Requirements:
    pip install requests openpyxl
"""

import argparse
import csv
import json
import os
import sys
import time

try:
    import requests
except ImportError:
    print("ERROR: 'requests' package required. Install with: pip install requests")
    sys.exit(1)

BASE_URL = "https://api.hubapi.com"
BATCH_CREATE_URL = f"{BASE_URL}/crm/v3/objects/contacts/batch/create"
SEARCH_URL = f"{BASE_URL}/crm/v3/objects/contacts/search"
DEFAULT_BATCH_SIZE = 100  # HubSpot max is 100
RATE_LIMIT_DELAY = 0.12

# Common column name -> HubSpot property mapping
AUTO_MAPPING = {
    # English
    "email": "email",
    "e-mail": "email",
    "email address": "email",
    "first name": "firstname",
    "firstname": "firstname",
    "first_name": "firstname",
    "last name": "lastname",
    "lastname": "lastname",
    "last_name": "lastname",
    "company": "company",
    "company name": "company",
    "phone": "phone",
    "phone number": "phone",
    "mobile": "mobilephone",
    "mobile phone": "mobilephone",
    "job title": "jobtitle",
    "title": "jobtitle",
    "website": "website",
    "city": "city",
    "state": "state",
    "country": "country",
    "zip": "zip",
    "zip code": "zip",
    "address": "address",
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    # French
    "prénom": "firstname",
    "prenom": "firstname",
    "nom": "lastname",
    "nom de famille": "lastname",
    "société": "company",
    "societe": "company",
    "entreprise": "company",
    "téléphone": "phone",
    "telephone": "phone",
    "portable": "mobilephone",
    "poste": "jobtitle",
    "fonction": "jobtitle",
    "ville": "city",
    "pays": "country",
    "adresse": "address",
    "code postal": "zip",
    "site web": "website",
}

VALID_LIFECYCLE_STAGES = [
    "subscriber", "lead", "marketingqualifiedlead",
    "salesqualifiedlead", "opportunity", "customer",
    "evangelist", "other",
]


def get_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }


def read_file(filepath: str) -> list[dict]:
    """Read CSV or Excel file and return list of row dicts."""
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        return read_excel(filepath)
    elif ext in (".csv", ".tsv"):
        return read_csv(filepath, delimiter="\t" if ext == ".tsv" else ",")
    else:
        print(f"ERROR: Unsupported file format '{ext}'. Use .csv, .tsv, or .xlsx")
        sys.exit(1)


def read_csv(filepath: str, delimiter: str = ",") -> list[dict]:
    """Read CSV/TSV file."""
    rows = []
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]

    for encoding in encodings:
        try:
            with open(filepath, "r", encoding=encoding) as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                rows = [row for row in reader]
            break
        except UnicodeDecodeError:
            continue
    else:
        print(f"ERROR: Could not decode {filepath} with any supported encoding.")
        sys.exit(1)

    return rows


def read_excel(filepath: str) -> list[dict]:
    """Read Excel file (first sheet)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: 'openpyxl' package required for Excel files. "
              "Install with: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(val).strip() if val is not None else ""
        if any(v for v in row_dict.values()):  # Skip empty rows
            data.append(row_dict)

    return data


def build_mapping(raw_rows: list[dict], explicit_mapping: str = None) -> dict:
    """Build column-to-property mapping. Returns {source_col: hubspot_property}."""
    if not raw_rows:
        return {}

    source_columns = list(raw_rows[0].keys())
    mapping = {}

    # Apply explicit mapping first
    if explicit_mapping:
        for pair in explicit_mapping.split(","):
            pair = pair.strip()
            if "=" not in pair:
                print(f"WARNING: Invalid mapping '{pair}', expected 'Column=property'")
                continue
            src, dest = pair.split("=", 1)
            src = src.strip()
            dest = dest.strip()
            if src in source_columns:
                mapping[src] = dest
            else:
                # Try case-insensitive match
                matched = False
                for col in source_columns:
                    if col.lower() == src.lower():
                        mapping[col] = dest
                        matched = True
                        break
                if not matched:
                    print(f"WARNING: Column '{src}' not found in file. Available: "
                          f"{', '.join(source_columns)}")

    # Auto-map remaining columns
    for col in source_columns:
        if col in mapping:
            continue
        normalized = col.strip().lower()
        if normalized in AUTO_MAPPING:
            mapping[col] = AUTO_MAPPING[normalized]

    return mapping


def validate_contacts(rows: list[dict], mapping: dict) -> tuple[list[dict], list[dict]]:
    """Validate contacts and return (valid, invalid) lists."""
    valid = []
    invalid = []

    email_prop = None
    for src, dest in mapping.items():
        if dest == "email":
            email_prop = src
            break

    for i, row in enumerate(rows, 1):
        properties = {}
        for src, dest in mapping.items():
            val = row.get(src, "").strip()
            if val and val.lower() != "none":
                properties[dest] = val

        if not properties:
            invalid.append({"row": i, "data": row, "reason": "No mapped properties"})
            continue

        email = properties.get("email", "")
        if email and "@" not in email:
            invalid.append({"row": i, "data": row, "reason": f"Invalid email: {email}"})
            continue

        if not email and not (properties.get("firstname") and properties.get("lastname")):
            invalid.append({"row": i, "data": row,
                          "reason": "Need email or firstname+lastname"})
            continue

        valid.append({"properties": properties, "source_row": i})

    return valid, invalid


def deduplicate_by_email(contacts: list[dict]) -> list[dict]:
    """Remove duplicates based on email, keeping first occurrence."""
    seen_emails = set()
    unique = []
    dupes = 0

    for contact in contacts:
        email = contact["properties"].get("email", "").lower()
        if email:
            if email in seen_emails:
                dupes += 1
                continue
            seen_emails.add(email)
        unique.append(contact)

    if dupes:
        print(f"  Removed {dupes} duplicate(s) by email")
    return unique


def check_existing_contacts(token: str, emails: list[str]) -> set:
    """Check which emails already exist in HubSpot. Returns set of existing emails."""
    existing = set()
    batch_size = 100

    for i in range(0, len(emails), batch_size):
        batch = emails[i:i + batch_size]
        for email in batch:
            payload = {
                "filterGroups": [{
                    "filters": [{
                        "propertyName": "email",
                        "operator": "EQ",
                        "value": email,
                    }]
                }],
                "limit": 1,
            }
            resp = requests.post(SEARCH_URL, headers=get_headers(token), json=payload)
            if resp.status_code == 200:
                results = resp.json().get("results", [])
                if results:
                    existing.add(email.lower())
            time.sleep(RATE_LIMIT_DELAY)

    return existing


def batch_create_contacts(token: str, contacts: list[dict],
                          lifecycle_stage: str = None,
                          batch_size: int = DEFAULT_BATCH_SIZE,
                          dry_run: bool = False) -> dict:
    """Create contacts in HubSpot in batches."""
    results = {"created": [], "failed": [], "skipped": []}
    total = len(contacts)

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Creating {total} contacts "
          f"in batches of {batch_size}...")

    for batch_start in range(0, total, batch_size):
        batch = contacts[batch_start:batch_start + batch_size]
        batch_end = min(batch_start + batch_size, total)
        batch_num = (batch_start // batch_size) + 1

        # Add lifecycle stage if specified
        inputs = []
        for contact in batch:
            props = dict(contact["properties"])
            if lifecycle_stage:
                props["lifecyclestage"] = lifecycle_stage
            inputs.append({"properties": props})

        if dry_run:
            print(f"  Batch {batch_num}: Would create {len(batch)} contacts "
                  f"(rows {batch_start + 1}-{batch_end})")
            for c in batch:
                results["skipped"].append({
                    "row": c["source_row"],
                    "email": c["properties"].get("email", "N/A"),
                    "reason": "dry_run",
                })
            continue

        payload = {"inputs": inputs}
        resp = requests.post(BATCH_CREATE_URL, headers=get_headers(token),
                           json=payload)

        if resp.status_code in (200, 201):
            created = resp.json().get("results", [])
            print(f"  Batch {batch_num}: Created {len(created)} contacts")
            for c in created:
                results["created"].append({
                    "id": c.get("id"),
                    "email": c.get("properties", {}).get("email", ""),
                })
        elif resp.status_code == 429:
            retry_after = int(resp.headers.get("Retry-After", 10))
            print(f"  Batch {batch_num}: Rate limited, waiting {retry_after}s...")
            time.sleep(retry_after)
            # Retry once
            resp = requests.post(BATCH_CREATE_URL, headers=get_headers(token),
                               json=payload)
            if resp.status_code in (200, 201):
                created = resp.json().get("results", [])
                print(f"  Batch {batch_num} (retry): Created {len(created)} contacts")
                for c in created:
                    results["created"].append({
                        "id": c.get("id"),
                        "email": c.get("properties", {}).get("email", ""),
                    })
            else:
                print(f"  Batch {batch_num} (retry): FAILED HTTP {resp.status_code}")
                results["failed"].extend([{
                    "row": c["source_row"],
                    "email": c["properties"].get("email", "N/A"),
                    "error": resp.text[:200],
                } for c in batch])
        else:
            error_data = resp.json() if resp.text else {}
            error_msg = error_data.get("message", resp.text[:200])
            print(f"  Batch {batch_num}: FAILED HTTP {resp.status_code} - {error_msg}")
            results["failed"].extend([{
                "row": c["source_row"],
                "email": c["properties"].get("email", "N/A"),
                "error": error_msg,
            } for c in batch])

        time.sleep(RATE_LIMIT_DELAY)

    return results


def print_summary(results: dict, invalid: list):
    """Print import summary."""
    print("\n--- Import Summary ---")
    print(f"  Created:   {len(results['created'])}")
    print(f"  Failed:    {len(results['failed'])}")
    print(f"  Skipped:   {len(results['skipped'])}")
    print(f"  Invalid:   {len(invalid)}")

    if results["failed"]:
        print("\n  Failed contacts:")
        for f in results["failed"][:10]:
            print(f"    - Row {f['row']} ({f['email']}): {f['error']}")
        if len(results["failed"]) > 10:
            print(f"    ... and {len(results['failed']) - 10} more")

    if invalid:
        print(f"\n  Invalid rows ({len(invalid)}):")
        for inv in invalid[:5]:
            print(f"    - Row {inv['row']}: {inv['reason']}")
        if len(invalid) > 5:
            print(f"    ... and {len(invalid) - 5} more")


def main():
    parser = argparse.ArgumentParser(
        description="Import contacts from CSV/Excel into HubSpot CRM"
    )
    parser.add_argument("--token", required=True,
                        help="HubSpot private app access token")
    parser.add_argument("--file", required=True,
                        help="Path to CSV/TSV/Excel file")
    parser.add_argument("--mapping",
                        help="Column mapping: 'Col1=prop1,Col2=prop2'")
    parser.add_argument("--lifecycle-stage", choices=VALID_LIFECYCLE_STAGES,
                        help="Lifecycle stage for imported contacts")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE,
                        help=f"Batch size for creation (max {DEFAULT_BATCH_SIZE})")
    parser.add_argument("--skip-duplicates", action="store_true",
                        help="Check and skip contacts that already exist in HubSpot")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate and show what would happen without importing")
    parser.add_argument("--output-json",
                        help="Path to write results as JSON")

    args = parser.parse_args()

    # Validate file exists
    if not os.path.isfile(args.file):
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    batch_size = min(args.batch_size, DEFAULT_BATCH_SIZE)

    # Read file
    print(f"Reading {args.file}...")
    raw_rows = read_file(args.file)
    print(f"  Found {len(raw_rows)} rows")

    if not raw_rows:
        print("ERROR: File is empty or has no data rows.")
        sys.exit(1)

    # Build mapping
    mapping = build_mapping(raw_rows, args.mapping)
    if not mapping:
        print("ERROR: No columns could be mapped to HubSpot properties.")
        print(f"  File columns: {', '.join(raw_rows[0].keys())}")
        print("  Use --mapping to specify: 'ColumnName=hubspot_property'")
        sys.exit(1)

    print(f"\n  Column mapping:")
    for src, dest in mapping.items():
        print(f"    {src} -> {dest}")

    # Validate
    print(f"\nValidating contacts...")
    valid, invalid = validate_contacts(raw_rows, mapping)
    print(f"  Valid: {len(valid)}, Invalid: {len(invalid)}")

    if not valid:
        print("ERROR: No valid contacts found after validation.")
        print_summary({"created": [], "failed": [], "skipped": []}, invalid)
        sys.exit(1)

    # Deduplicate within file
    valid = deduplicate_by_email(valid)

    # Check existing contacts in HubSpot
    if args.skip_duplicates and not args.dry_run:
        emails = [c["properties"]["email"] for c in valid
                  if c["properties"].get("email")]
        if emails:
            print(f"\nChecking {len(emails)} emails against HubSpot...")
            existing = check_existing_contacts(args.token, emails)
            if existing:
                before = len(valid)
                valid = [c for c in valid
                        if c["properties"].get("email", "").lower() not in existing]
                print(f"  Skipped {before - len(valid)} existing contacts")

    # Import
    results = batch_create_contacts(
        token=args.token,
        contacts=valid,
        lifecycle_stage=args.lifecycle_stage,
        batch_size=batch_size,
        dry_run=args.dry_run,
    )

    print_summary(results, invalid)

    # Save results
    if args.output_json:
        output = {
            "results": results,
            "invalid_rows": [{"row": inv["row"], "reason": inv["reason"]}
                           for inv in invalid],
            "mapping_used": mapping,
        }
        with open(args.output_json, "w") as f:
            json.dump(output, f, indent=2)
        print(f"\nResults saved to {args.output_json}")

    if results["failed"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
